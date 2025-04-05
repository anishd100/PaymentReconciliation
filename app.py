import os
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, jsonify
from flask_mysqldb import MySQL




import MySQLdb.cursors  # Import DictCursor


app = Flask(__name__)

# MySQL Configuration
app.config['MYSQL_HOST'] = os.environ.get('MYSQL_HOST', 'mysql')
app.config['MYSQL_USER'] = os.environ.get('MYSQL_USER', 'root')
app.config['MYSQL_PASSWORD'] = os.environ.get('MYSQL_PASSWORD', 'admin')
app.config['MYSQL_DB'] = os.environ.get('MYSQL_DB', 'managementdb')

# Initialize MySQL
mysql = MySQL(app)

def init_db():
    with app.app_context():
        connection = mysql.connection  # ✅ Corrected
        cursor = connection.cursor()  # ✅ Now correctly using cursor

        try:
            cursor.execute('USE managementdb')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS states (
                    State_Id int NOT NULL AUTO_INCREMENT,
                    State_Name varchar(255) NOT NULL,
                    PRIMARY KEY (State_Id)
                );
            ''')

            connection.commit()
        except Exception as e:
            connection.rollback()
            print(f"Database initialization error: {e}")
        finally:
            cursor.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add_state', methods=['GET', 'POST'])
def add_state():
    connection = mysql.connection  # ✅ Fix
    cur = connection.cursor()

    if request.method == 'POST':
        state_name = request.form['state_Name'].strip()

        try:
            cur.execute("SELECT * FROM states WHERE State_Name = %s", (state_name,))
            existing_state = cur.fetchone()

            if existing_state:
                return jsonify({"status": "error", "message": "State already exists!"})

            cur.execute("INSERT INTO states (State_Name) VALUES (%s)", (state_name,))
            connection.commit()
            return redirect(url_for('add_state'))
        except Exception as e:
            print(f"Error inserting data: {e}")
            return jsonify({"status": "error", "message": "Failed to add state!"})
        finally:
            cur.close()

    try:
        cur.execute("SELECT * FROM states")
        statedata = cur.fetchall()
    except Exception as e:
        print(f"Error retrieving data: {e}")
        statedata = []
    finally:
        cur.close()

    return render_template('add_state.html', statedata=statedata)

@app.route('/delete_state/<int:id>', methods=['GET'])
def deleteState(id):
    connection = mysql.connection  # ✅ Fix
    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM states WHERE State_ID = %s", (id,))
        connection.commit()
    except Exception as e:
        print(f"Error deleting data: {e}")
        return "Failed to delete state", 500
    finally:
        cursor.close()

    return redirect(url_for('add_state'))

@app.route('/edit_state/<int:id>', methods=['GET', 'POST'])
def editState(id):
    connection = mysql.connection  # ✅ Fix
    cursor = connection.cursor()

    if request.method == 'POST':
        state_name = request.form['state_Name'].strip()
        try:
            cursor.execute("UPDATE states SET State_Name = %s WHERE State_ID = %s", (state_name, id))
            connection.commit()
            return redirect(url_for('add_state'))
        except Exception as e:
            print(f"Error updating data: {e}")
            return "Failed to update state", 500
        finally:
            cursor.close()

    try:
        cursor.execute("SELECT * FROM states WHERE State_ID = %s", (id,))
        state = cursor.fetchone()
        if state is None:
            return "State not found", 404
    except Exception as e:
        print(f"Error retrieving data: {e}")
        return "Failed to retrieve state", 500
    finally:
        cursor.close()

    return render_template('edit_state.html', state=state)

@app.route('/check_state', methods=['POST'])
def check_state():
    state_name = request.form['state_Name'].strip()
    connection = mysql.connection  # ✅ Fix
    cursor = connection.cursor()

    try:
        cursor.execute("SELECT * FROM states WHERE State_Name = %s", (state_name,))
        existing_state = cursor.fetchone()

        if existing_state:
            return jsonify({"exists": True, "message": "State already exists!"})
        else:
            return jsonify({"exists": False})
    except Exception as e:
        print(f"Error checking data: {e}")
        return jsonify({"exists": False, "message": "Error checking state!"})
    finally:
        cursor.close()
        
        ##District
        
import MySQLdb.cursors  # Import DictCursor
@app.route('/add_district', methods=['GET', 'POST'])
def add_district():
    connection = mysql.connection
    if not connection:
        print("❌ Database connection failed!")
        return "Database connection failed!", 500  # Stop execution if connection is invalid
    
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    districtdata = []
    states = []
    message = ""

    try:
        # ✅ Fetch all states for dropdown
        cur.execute("SELECT State_ID, State_Name FROM states")
        states = cur.fetchall()
        print("✅ Fetched states:", states)  # Debugging

        # ✅ Fetch all districts
        cur.execute("""
            SELECT d.District_id, d.District_Name, s.State_Name, s.State_ID 
            FROM districts d 
            JOIN states s ON d.State_Id = s.State_ID
        """)
        districtdata = cur.fetchall()

        # ✅ Handle POST request to insert new district
        if request.method == 'POST':
            district_name = request.form.get('district_Name', '').strip()
            state_id = request.form.get('state_Id', '').strip()

            print(f"📝 Received - District: {district_name}, State ID: {state_id}")  # Debugging

            if not state_id or not state_id.isdigit():
                message = "❌ Please select a valid state."
            elif not district_name or not district_name.replace(" ", "").isalpha():
                message = "❌ Invalid District Name! Only letters allowed."
            else:
                state_id = int(state_id)

                # ✅ Check if the state ID exists before inserting
                cur.execute("SELECT * FROM states WHERE State_ID = %s", (state_id,))
                state_exists = cur.fetchone()
                if not state_exists:
                    message = "❌ Selected State does not exist!"
                    return render_template('add_district.html', states=states, districtdata=districtdata, message=message)

                # ✅ Check if district already exists
                cur.execute("SELECT * FROM districts WHERE District_Name = %s AND State_Id = %s",
                            (district_name, state_id))
                existing_district = cur.fetchone()

                if existing_district:
                    message = "⚠️ District already exists!"
                else:
                    print(f"📝 Inserting - District: {district_name}, State ID: {state_id}")  # Debugging

                    try:
                        cur.execute("INSERT INTO districts (District_Name, State_Id) VALUES (%s, %s)",
                                    (district_name, state_id))
                        connection.commit()  # ✅ Commit transaction
                        print("✅ District inserted successfully!")  # Debugging
                        return redirect(url_for('add_district'))  # 🔄 Refresh page

                    except MySQLdb.Error as e:
                        connection.rollback()
                        print(f"❌ MySQL Error: {e}")
                        message = f"Failed to add district! {str(e)}"

    except MySQLdb.Error as e:
        print(f"❌ MySQL Database Error: {e}")
        message = f"Database Error: {e}"
    except Exception as e:
        print(f"❌ Unexpected Error: {e}")
        message = f"Unexpected Error: {e}"
    finally:
        cur.close()

    return render_template('add_district.html', states=states, districtdata=districtdata, message=message)



# Delete District
@app.route('/delete_district/<int:district_id>', methods=['GET'])
def delete_district(district_id):
    connection = mysql.connection
    cur = connection.cursor()

    try:
        cur.execute("DELETE FROM districts WHERE District_id = %s", (district_id,))
        connection.commit()
        print(f"✅ Deleted district with ID {district_id}")  # Debugging
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting district: {e}")
        return "Failed to delete district", 500
    finally:
        cur.close()

    return redirect(url_for('add_district'))


# AJAX Route to Check District Existence
@app.route('/check_district', methods=['POST'])
def check_district():
    district_name = request.form['district_Name'].strip()
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)

    try:
        cur.execute('SELECT * FROM districts WHERE District_Name = %s', (district_name,))
        existing_district = cur.fetchone()
        return jsonify({'exists': existing_district is not None})
    except Exception as e:
        print(f"❌ Error checking district: {e}")
        return jsonify({"exists": False, "message": "Database query failed"}), 500
    finally:
        cur.close()


# Edit District
@app.route('/edit_district/<int:district_id>', methods=['GET', 'POST'])
def edit_district(district_id):
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    district = None
    states = []
    message = ""

    try:
        # ✅ Fetch district details (Ensure State_Id is selected)
        cur.execute('SELECT District_id, District_Name, State_Id FROM districts WHERE District_id = %s', (district_id,))
        district = cur.fetchone()

        if not district:
            return "District not found", 404

        # ✅ Fetch states for dropdown
        cur.execute('SELECT State_ID, State_Name FROM states')
        states = cur.fetchall()

        if request.method == 'POST':
            new_name = request.form['district_Name'].strip()
            state_id = request.form['state_Id']

            if not new_name.isalpha():
                message = "Invalid District Name! Only letters allowed."
            else:
                cur.execute(
                    'SELECT * FROM districts WHERE District_Name = %s AND State_Id = %s AND District_id != %s',
                    (new_name, state_id, district_id))
                existing_district = cur.fetchone()

                if existing_district:
                    message = "District already exists!"
                else:
                    cur.execute('UPDATE districts SET District_Name = %s, State_Id = %s WHERE District_id = %s',
                                (new_name, state_id, district_id))
                    connection.commit()
                    print("✅ District updated successfully!")  # Debugging
                    return redirect(url_for('add_district'))

    except Exception as e:
        connection.rollback()
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_district.html', district=district, states=states, message=message)





#Blocks
import MySQLdb.cursors  # Import DictCursor

@app.route('/add_block', methods=['GET', 'POST'])
def add_new_block():  # 🔥 Renamed to avoid duplicate issue
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    block_data = []
    states = []
    districts = []
    message = ""

    try:
        # ✅ Retrieve all states for dropdown
        cur.execute("SELECT State_ID, State_Name FROM states")
        states = cur.fetchall()

        # ✅ Handle POST request to add a block
        if request.method == 'POST':
            block_name = request.form.get('block_Name', '').strip()
            district_id = request.form.get('district_Id', '').strip()

            print(f"📝 Received: Block Name: {block_name}, District ID: {district_id}")  # Debugging

            if not block_name or not block_name.replace(" ", "").isalpha():
                message = "Invalid Block Name! Only letters allowed."
            elif not district_id or not district_id.isdigit():
                message = "Please select a valid district."
            else:
                district_id = int(district_id)

                try:
                    cur.execute("INSERT INTO blocks (Block_Name, District_id) VALUES (%s, %s)",
                                (block_name, district_id))
                    connection.commit()
                    print("✅ Block added successfully!")
                    return redirect(url_for('add_new_block'))  # 🔥 Updated redirect
                except Exception as e:
                    connection.rollback()
                    print(f"❌ Error inserting block: {e}")
                    message = "Failed to add block!"

        # ✅ Retrieve all blocks
        cur.execute("""
            SELECT b.Block_Id, b.Block_Name, d.District_Name, COALESCE(s.State_Name, 'Unknown') AS State_Name
            FROM blocks b
            JOIN districts d ON b.District_id = d.District_id
            LEFT JOIN states s ON d.State_Id = s.State_ID
        """)
        block_data = cur.fetchall()

        # ✅ Retrieve districts for dropdown
        cur.execute("SELECT District_id, District_Name FROM districts")
        districts = cur.fetchall()

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = f"Database Error: {e}"
    finally:
        cur.close()

    return render_template('add_block.html', block_data=block_data, states=states, districts=districts, message=message)

#-----------------------------edit block

@app.route('/edit_block/<int:block_id>', methods=['GET', 'POST'])
def edit_block(block_id):
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    block_data = None
    states = []
    districts = []
    message = ""

    try:
        # ✅ Retrieve block details (including Block_Id & State_Id)
        cur.execute("""
            SELECT b.Block_Id, b.Block_Name, b.District_id, d.State_Id 
            FROM blocks b
            JOIN districts d ON b.District_id = d.District_id
            WHERE b.Block_Id = %s
        """, (block_id,))
        block_data = cur.fetchone()

        if not block_data:
            return "Block not found", 404  # ✅ Handle missing block

        # ✅ Retrieve all states
        cur.execute("SELECT State_ID, State_Name FROM states")
        states = cur.fetchall()

        # ✅ Retrieve all districts for the selected state
        cur.execute("SELECT District_id, District_Name FROM districts WHERE State_Id = %s", (block_data['State_Id'],))
        districts = cur.fetchall()

        if request.method == 'POST':
            block_name = request.form['block_Name'].strip()
            district_id = int(request.form['district_Id'])

            if not block_name.replace(" ", "").isalpha():
                message = "Invalid Block Name! Only letters and spaces allowed."
            else:
                # ✅ Check if block name already exists in the district
                cur.execute(
                    "SELECT * FROM blocks WHERE Block_Name = %s AND District_id = %s AND Block_Id != %s",
                    (block_name, district_id, block_id))
                existing_block = cur.fetchone()

                if existing_block:
                    message = "Block already exists!"
                else:
                    # ✅ Update block data
                    cur.execute("UPDATE blocks SET Block_Name = %s, District_id = %s WHERE Block_Id = %s",
                                (block_name, district_id, block_id))
                    connection.commit()
                    print("✅ Block updated successfully!")  # Debugging

                    # ✅ Fetch the updated block details
                    cur.execute("""
                        SELECT b.Block_Id, b.Block_Name, b.District_id, d.State_Id 
                        FROM blocks b
                        JOIN districts d ON b.District_id = d.District_id
                        WHERE b.Block_Id = %s
                    """, (block_id,))
                    block_data = cur.fetchone()

                    # ✅ Re-fetch districts in case the state changed
                    cur.execute("SELECT District_id, District_Name FROM districts WHERE State_Id = %s", (block_data['State_Id'],))
                    districts = cur.fetchall()

                    message = "Block updated successfully!"

    except Exception as e:
        connection.rollback()
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_block.html', block_data=block_data, states=states, districts=districts, message=message)


@app.route('/delete_block/<int:block_id>', methods=['GET'])
def delete_block(block_id):
    connection = mysql.connection
    cur = connection.cursor()
    try:
        cur.execute("DELETE FROM blocks WHERE Block_Id = %s", (block_id,))
        connection.commit()
        print(f"✅ Deleted block with ID {block_id}")  # Debugging
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting block: {e}")
        return "Failed to delete block", 500
    finally:
        cur.close()

    return redirect(url_for('add_block'))


# ✅ Get Districts by State ID (AJAX Request)
@app.route('/get_districts/<int:state_id>', methods=['GET'])
def get_districts(state_id):
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    districts = []

    try:
        cur.execute("SELECT District_id, District_Name FROM districts WHERE State_Id = %s", (state_id,))
        districts = cur.fetchall()
    except Exception as e:
        print(f"❌ Error fetching districts: {e}")
        return jsonify({"error": "Failed to fetch districts"}), 500
    finally:
        cur.close()

    return jsonify({"districts": districts})



# ------------------------- Village controller ------------------------------------------

# Add Village
@app.route('/add_village', methods=['GET', 'POST'])
def add_village():
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Fix
    states = []
    villages = []
    message = ""

    try:
        # Fetch states for dropdown
        cur.execute("SELECT State_ID, State_Name FROM states")
        states = cur.fetchall()

        # Fetch all villages
        cur.execute("SELECT * FROM villages")
        villages = cur.fetchall()

        if request.method == 'POST':
            block_id = request.form.get('block_Id', '').strip()
            village_name = request.form.get('Village_Name', '').strip()

            if not village_name or not village_name.replace(" ", "").isalpha():
                message = "Invalid Village Name! Only letters allowed."
            elif not block_id or not block_id.isdigit():
                message = "Please select a valid block."
            else:
                block_id = int(block_id)

                # Insert village into database
                try:
                    cur.execute("INSERT INTO villages (Village_Name, Block_Id) VALUES (%s, %s)",
                                (village_name, block_id))
                    connection.commit()
                    print("✅ Village inserted successfully!")  # Debugging
                    return redirect(url_for('add_village'))
                except Exception as e:
                    connection.rollback()
                    print(f"❌ Error inserting village: {e}")
                    message = "Failed to add village!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = f"Database Error: {e}"
    finally:
        cur.close()

    return render_template('add_village.html', states=states, villages=villages, message=message)


# Get Blocks by District ID (AJAX Request)
@app.route('/get_blocks/<int:district_id>', methods=['GET'])
def get_blocks(district_id):
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    blocks = []

    try:
        cur.execute("SELECT Block_id, Block_Name FROM blocks WHERE District_id = %s", (district_id,))
        blocks = cur.fetchall()
    except Exception as e:
        print(f"❌ Error fetching blocks: {e}")
        return jsonify({"error": "Failed to fetch blocks"}), 500
    finally:
        cur.close()

    return jsonify({"blocks": blocks})


# Edit Village
@app.route('/edit_village/<int:village_id>', methods=['GET', 'POST'])
def edit_village(village_id):
    connection = mysql.connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)
    village = None
    states = []
    message = ""

    try:
        # ✅ Fetch village details (Ensure Block_Id is retrieved)
        cur.execute("SELECT Village_Id, Village_Name, Block_Id FROM villages WHERE Village_Id = %s", (village_id,))
        village = cur.fetchone()

        if not village:
            return "Village not found", 404

        # ✅ Fetch all states
        cur.execute("SELECT State_ID, State_Name FROM states")
        states = cur.fetchall()

        if request.method == 'POST':
            village_name = request.form['Village_Name'].strip()
            block_id = request.form['block_Id']

            if not village_name.isalpha():
                message = "Invalid Village Name! Only letters allowed."
            else:
                cur.execute(
                    "UPDATE villages SET Village_Name = %s, Block_Id = %s WHERE Village_Id = %s",
                    (village_name, block_id, village_id))
                connection.commit()
                print("✅ Village updated successfully!")  # Debugging
                return redirect(url_for('add_village'))

    except Exception as e:
        connection.rollback()
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_village.html', village=village, states=states, message=message)



# Delete Village
@app.route('/delete_village/<int:village_id>', methods=['GET'])
def delete_village(village_id):
    connection = mysql.connection
    cur = connection.cursor()

    try:
        cur.execute("DELETE FROM villages WHERE Village_Id = %s", (village_id,))
        connection.commit()
        print(f"✅ Deleted village with ID {village_id}")  # Debugging
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting village: {e}")
        return "Failed to delete village", 500
    finally:
        cur.close()

    return redirect(url_for('add_village'))

# -------------------------------- Invoice controller ------------------------------------------
# this is invoice page to add data
@app.route('/add_invoice', methods=['GET', 'POST'])
def add_invoice():
    if request.method == 'POST':
        # Collecting form data
        pmc_no = request.form['pmc_no']
        village_id = request.form['Village_Id']
        work_Type = request.form['Work_Type']
        invoice_details = request.form['invoice_details']
        invoice_date = request.form['invoice_date']
        invoice_no = request.form['invoice_no']
        basic_amount = float(request.form['basic_amount'])
        debit_amount = float(request.form['debit_amount'])
        after_debit_amount = float(request.form['after_debit_amount'])
        amount = float(request.form['Amount'])

        gst_amount = float(request.form['gst_amount'])
        tds_amount = float(request.form['tds_amount'])
        sd_amount = float(request.form['sd_amount'])

        on_commission = float(request.form['On_Commission'])
        hydro_testing = float(request.form['Hydro_Testing'])

        hold_amount = float(request.form['hold_amount'])
        gst_sd_amount = float(request.form['gst_sd_amount'])
        final_amount = float(request.form['final_amount'])

        # Inserting data into MySQL
        connection = mysql.connection.cursor()
        cursor = connection.cursor()

        cursor.execute(
            "INSERT INTO Invoice (PMC_No,Village_Id , Work_Type,Invoice_Details, Invoice_Date, Invoice_No,Basic_Amount,Debit_Amount, After_Debit_Amount, Amount,GST_Amount, TDS_Amount, SD_Amount,On_Commission,Hydro_Testing, Hold_Amount, GST_SD_Amount, Final_Amount) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s,%s)",
            (pmc_no, village_id, work_Type, invoice_details, invoice_date, invoice_no, basic_amount,
             debit_amount, after_debit_amount, amount, gst_amount, tds_amount, sd_amount, on_commission, hydro_testing,
             hold_amount, gst_sd_amount, final_amount))

        connection.commit()
        cursor.execute("select * from Invoice")
        data=cursor.fetchall()
        print(data,flush=True)
        cursor.close()
        connection.close()
        return redirect(url_for('add_invoice'))

    return render_template('add_invoice.html')


# ----------------------------- Payment controller ------------------------------------------
# this is Payment Page to add data
import MySQLdb.cursors  # Import DictCursor

@app.route('/add_payment', methods=['GET', 'POST'])
def add_payment():
    connection = mysql.connection  # ✅ Corrected connection initialization
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    invoices = []  # List to hold invoices for the dropdown
    payments = []  # List to hold payment history
    message = ""

    try:
        # ✅ Retrieve all invoices for dropdown
        cur.execute("SELECT Invoice_Id, Invoice_details, PMC_No FROM invoice")
        invoices = cur.fetchall()

        # ✅ Retrieve payment history
        cur.execute("SELECT Payment_Id, PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR FROM payment")
        payments = cur.fetchall()

        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            amount = request.form['amount']
            tds_amount = request.form['tds_amount']
            total_amount = request.form['total_amount']
            utr = request.form['utr']

            print(f"📝 Received Data: Invoice No: {invoice_no}, Amount: {amount}, TDS: {tds_amount}, Total: {total_amount}, UTR: {utr}")

            try:
                cur.execute('''
                    INSERT INTO payment (PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR)
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (pmc_no, invoice_no, amount, tds_amount, total_amount, utr))

                connection.commit()
                print("✅ Payment added successfully!")
                return redirect(url_for('add_payment'))  # Refresh the page

            except Exception as e:
                connection.rollback()
                print(f"❌ Error inserting payment: {e}")
                message = "Failed to add payment!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('add_payment.html', invoices=invoices, payments=payments, message=message)


@app.route('/edit_payment/<int:payment_id>', methods=['GET', 'POST'])
def edit_payment(payment_id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    payment_data = {}
    invoices = []
    message = ""

    try:
        # ✅ Fetch all invoices for dropdown
        cur.execute("SELECT Invoice_Id, Invoice_details FROM invoice")
        invoices = cur.fetchall()

        # ✅ Fetch payment data
        cur.execute("""
            SELECT Payment_Id, PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR 
            FROM payment 
            WHERE Payment_Id = %s
        """, (payment_id,))
        payment_data = cur.fetchone()

        if not payment_data:
            return "Payment not found", 404

        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            amount = request.form['amount']
            tds_amount = request.form['tds_amount']
            total_amount = request.form['total_amount']
            utr = request.form['utr']

            print(f"📝 Updating Payment ID {payment_id}: {pmc_no}, {invoice_no}, {amount}, {tds_amount}, {total_amount}, {utr}")

            try:
                cur.execute('''
                    UPDATE payment 
                    SET PMC_No=%s, Invoice_No=%s, Amount=%s, TDS_Amount=%s, Total_Amount=%s, UTR=%s
                    WHERE Payment_Id = %s
                ''', (pmc_no, invoice_no, amount, tds_amount, total_amount, utr, payment_id))

                connection.commit()
                print("✅ Payment updated successfully!")
                return redirect(url_for('add_payment'))

            except Exception as e:
                connection.rollback()
                print(f"❌ Error updating payment: {e}")
                message = "Failed to update payment!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_payment.html', payment_data=payment_data, invoices=invoices, message=message)


@app.route('/delete_payment/<int:payment_id>', methods=['GET'])
def delete_payment(payment_id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor()

    try:
        print(f"🗑 Deleting Payment ID {payment_id}")  # Debugging
        cur.execute("DELETE FROM payment WHERE Payment_Id = %s", (payment_id,))
        connection.commit()
        print("✅ Payment deleted successfully!")
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting payment: {e}")
        return "Failed to delete payment", 500
    finally:
        cur.close()

    return redirect(url_for('add_payment'))



# ------------------------- GST Release controller ------------------------------------------
import MySQLdb.cursors  # Import DictCursor

@app.route('/add_gst_release', methods=['GET', 'POST'])
def add_gst_release():
    connection = mysql.connection  # ✅ Corrected connection initialization
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    gst_releases = []  # List to hold GST Release history
    invoices = []  # List to hold invoices for dropdown
    message = ""

    try:
        # ✅ Retrieve all invoices for dropdown
        cur.execute("SELECT Invoice_Id, Invoice_details FROM invoice")
        invoices = cur.fetchall()

        # ✅ Retrieve GST Release history
        cur.execute("SELECT GST_Release_Id, PMC_No, Invoice_No, Basic_Amount, Final_Amount FROM gst_release")
        gst_releases = cur.fetchall()

        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            basic_amount = request.form['basic_amount']
            final_amount = request.form['final_amount']

            print(f"📝 Received Data: PMC No: {pmc_no}, Invoice No: {invoice_no}, Basic: {basic_amount}, Final: {final_amount}")

            try:
                cur.execute('''
                    INSERT INTO gst_release (PMC_No, Invoice_No, Basic_Amount, Final_Amount)
                    VALUES (%s, %s, %s, %s)
                ''', (pmc_no, invoice_no, basic_amount, final_amount))

                connection.commit()
                print("✅ GST Release added successfully!")
                return redirect(url_for('add_gst_release'))  # Refresh the page

            except Exception as e:
                connection.rollback()
                print(f"❌ Error inserting GST Release: {e}")
                message = "Failed to add GST Release!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('add_gst_release.html', invoices=invoices, gst_releases=gst_releases, message=message)


@app.route('/edit_gst_release/<int:gst_release_id>', methods=['GET', 'POST'])
def edit_gst_release(gst_release_id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    gst_release_data = {}
    invoices = []
    message = ""

    try:
        # ✅ Fetch all invoices for dropdown
        cur.execute("SELECT Invoice_Id, Invoice_details FROM invoice")
        invoices = cur.fetchall()

        # ✅ Fetch GST Release data
        cur.execute("""
            SELECT GST_Release_Id, PMC_No, Invoice_No, Basic_Amount, Final_Amount 
            FROM gst_release 
            WHERE GST_Release_Id = %s
        """, (gst_release_id,))
        gst_release_data = cur.fetchone()

        if not gst_release_data:
            return "GST Release not found", 404

        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            basic_amount = request.form['basic_amount']
            final_amount = request.form['final_amount']

            print(f"📝 Updating GST Release ID {gst_release_id}: {pmc_no}, {invoice_no}, {basic_amount}, {final_amount}")

            try:
                cur.execute('''
                    UPDATE gst_release 
                    SET PMC_No = %s, Invoice_No = %s, Basic_Amount = %s, Final_Amount = %s
                    WHERE GST_Release_Id = %s
                ''', (pmc_no, invoice_no, basic_amount, final_amount, gst_release_id))

                connection.commit()
                print("✅ GST Release updated successfully!")
                return redirect(url_for('add_gst_release'))

            except Exception as e:
                connection.rollback()
                print(f"❌ Error updating GST Release: {e}")
                message = "Failed to update GST Release!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_gst_release.html', gst_release_data=gst_release_data, invoices=invoices, message=message)


@app.route('/delete_gst_release/<int:gst_release_id>', methods=['GET'])
def delete_gst_release(gst_release_id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor()

    try:
        print(f"🗑 Deleting GST Release ID {gst_release_id}")  # Debugging
        cur.execute("DELETE FROM gst_release WHERE GST_Release_Id = %s", (gst_release_id,))
        connection.commit()
        print("✅ GST Release deleted successfully!")
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting GST Release: {e}")
        return "Failed to delete GST Release", 500
    finally:
        cur.close()

    return redirect(url_for('add_gst_release'))



# ------------------------- Subcontractor controller ------------------------------------------
# this is Sub-Contractor Page to show data or  Add data
import MySQLdb.cursors  # Import DictCursor

@app.route('/subcontractor', methods=['GET', 'POST'])
def subcontract():
    connection = mysql.connection  # ✅ Corrected connection initialization
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    subcontractors = []  # List to store retrieved subcontractors
    message = ""

    try:
        # ✅ Fetch existing subcontractors
        cur.execute('SELECT * FROM subcontractors;')
        subcontractors = cur.fetchall()

        if request.method == 'POST':
            contractor_data = {
                'Contractor_Name': request.form['Contractor_Name'],
                'Address': request.form['Address'],
                'Mobile_No': request.form['Mobile_No'],
                'PAN_No': request.form['PAN_No'],
                'Email': request.form['Email'],
                'Gender': request.form['Gender'],
                'GST_Registration_Type': request.form['GST_Registration_Type'],
                'GST_No': request.form['GST_No'],
                'Contractor_password': request.form['Contractor_password'],
            }

            print(f"📝 Received Contractor Data: {contractor_data}")  # Debugging

            try:
                # ✅ Insert new subcontractor
                cur.execute("""
                    INSERT INTO subcontractors 
                    (Contractor_Name, Address, Mobile_No, PAN_No, Email, Gender, 
                     GST_Registration_Type, GST_No, Contractor_password) 
                    VALUES (%(Contractor_Name)s, %(Address)s, %(Mobile_No)s, %(PAN_No)s, %(Email)s, 
                            %(Gender)s, %(GST_Registration_Type)s, %(GST_No)s, %(Contractor_password)s)
                """, contractor_data)
                connection.commit()
                print("✅ Subcontractor added successfully!")

                return redirect(url_for('subcontract'))  # Refresh page

            except Exception as e:
                connection.rollback()
                print(f"❌ Error inserting subcontractor: {e}")
                message = "Failed to add subcontractor!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('add_subcontractor.html', subcontractors=subcontractors, message=message)


@app.route('/edit_subcontractor/<int:id>', methods=['GET', 'POST'])
def edit_subcontractor(id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)  # ✅ Used DictCursor
    subcontractor = None
    message = ""

    try:
        # ✅ Fetch subcontractor details
        cur.execute('SELECT * FROM subcontractors WHERE Contractor_Id = %s', (id,))
        subcontractor = cur.fetchone()

        if not subcontractor:
            return "Subcontractor not found", 404

        if request.method == 'POST':
            updated_data = {
                'Contractor_Name': request.form['Contractor_Name'],
                'Address': request.form['Address'],
                'Mobile_No': request.form['Mobile_No'],
                'PAN_No': request.form['PAN_No'],
                'Email': request.form['Email'],
                'Gender': request.form['Gender'],
                'GST_Registration_Type': request.form['GST_Registration_Type'],
                'GST_No': request.form['GST_No'],
                'Contractor_password': request.form['Contractor_password'],
                'id': id
            }

            print(f"📝 Updating Contractor ID {id}: {updated_data}")  # Debugging

            try:
                cur.execute("""
                    UPDATE subcontractors 
                    SET Contractor_Name=%(Contractor_Name)s, Address=%(Address)s, 
                        Mobile_No=%(Mobile_No)s, PAN_No=%(PAN_No)s, Email=%(Email)s, 
                        Gender=%(Gender)s, GST_Registration_Type=%(GST_Registration_Type)s, 
                        GST_No=%(GST_No)s, Contractor_password=%(Contractor_password)s
                    WHERE Contractor_Id=%(id)s
                """, updated_data)

                connection.commit()
                print("✅ Subcontractor updated successfully!")
                return redirect(url_for('subcontract'))

            except Exception as e:
                connection.rollback()
                print(f"❌ Error updating subcontractor: {e}")
                message = "Failed to update subcontractor!"

    except Exception as e:
        print(f"❌ Database Error: {e}")
        message = "Database Error!"
    finally:
        cur.close()

    return render_template('edit_subcontractor.html', subcontractor=subcontractor, message=message)


@app.route('/delete_subcontractor/<int:id>', methods=['GET'])
def delete_subcontractor(id):
    connection = mysql.connection  # ✅ Fix connection
    cur = connection.cursor()

    try:
        print(f"🗑 Deleting Subcontractor ID {id}")  # Debugging
        cur.execute("DELETE FROM subcontractors WHERE Contractor_Id = %s", (id,))
        connection.commit()
        print("✅ Subcontractor deleted successfully!")
    except Exception as e:
        connection.rollback()
        print(f"❌ Error deleting subcontractor: {e}")
        return "Failed to delete subcontractor", 500
    finally:
        cur.close()

    return redirect(url_for('subcontract'))


# ------------------ upload Xcel file -----------------------------
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# ----------------- Upload Excel file html page url -------------------------------------------------
@app.route('/upload_excel_file', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            return redirect(url_for('show_table', filename=file.filename))
    return render_template('upload_excel_file.html')


# -------------------------- Show Excel file data in html table  -------------------------------------------------------------------------------------

import MySQLdb.cursors  # ✅ Import DictCursor

@app.route('/show_table/<filename>')
def show_table(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    file_info = {
        "Subcontractor": sheet.cell(row=1, column=2).value,
        "State": sheet.cell(row=2, column=2).value,
        "District": sheet.cell(row=3, column=2).value,
        "Block": sheet.cell(row=4, column=2).value,
    }

    errors = []
    subcontractor_data = None
    state_data = None
    district_data = None
    block_data = None

    # ✅ Get a fresh database connection
    connection = mysql.connection
    try:
        connection.ping(reconnect=True)  # ✅ Ensure connection is active
        cursor = connection.cursor(cursorclass=MySQLdb.cursors.DictCursor)

        # Validate State
        cursor.execute("SELECT State_ID, State_Name FROM states WHERE State_Name = %s", (file_info['State'],))
        state_data = cursor.fetchone()
        if not state_data:
            errors.append(f"State '{file_info['State']}' is not valid. Please add it.")

        if state_data:
            cursor.execute(
                "SELECT District_ID, District_Name FROM districts WHERE District_Name = %s AND State_ID = %s",
                (file_info['District'], state_data['State_ID']))
            district_data = cursor.fetchone()
            if not district_data:
                errors.append(f"District '{file_info['District']}' is not valid under state '{file_info['State']}'.")

        if district_data:
            cursor.execute(
                "SELECT Block_Id, Block_Name FROM blocks WHERE Block_Name = %s AND District_ID = %s",
                (file_info['Block'], district_data['District_ID']))
            block_data = cursor.fetchone()
            if not block_data:
                errors.append(f"Block '{file_info['Block']}' is not valid under district '{file_info['District']}'.")

        # Validate Subcontractor
        cursor.execute("SELECT Contractor_Id, Contractor_Name FROM SubContractors WHERE Contractor_Name = %s",
                       (file_info['Subcontractor'],))
        subcontractor_data = cursor.fetchone()
        if not subcontractor_data:
            # Add subcontractor if not found
            cursor.execute("INSERT INTO subcontractors (Contractor_Name) VALUES (%s)", (file_info['Subcontractor'],))
            connection.commit()
            cursor.execute("SELECT Contractor_Id, Contractor_Name FROM SubContractors WHERE Contractor_Name = %s",
                           (file_info['Subcontractor'],))
            subcontractor_data = cursor.fetchone()

    except MySQLdb.OperationalError as e:
        print(f"⚠️ MySQL error: {e}")
        return "Database operation failed", 500
    finally:
        cursor.close()
        connection.close()  # ✅ Ensure connection is properly closed

    # Extract variables and data from the Excel file
    variables = [
        sheet.cell(row=5, column=j).value
        for j in range(1, sheet.max_column + 1)
        if sheet.cell(row=5, column=j).value is not None
    ]

    data = [
        [
            sheet.cell(row=i, column=j).value
            for j in range(1, len(variables) + 1)
        ]
        for i in range(6, sheet.max_row + 1)
        if sheet.cell(row=i, column=1).value is not None
           and sum(1 for j in range(1, len(variables) + 1) if sheet.cell(row=i, column=j).value is not None) >= 4
    ]

    return render_template(
        'display_excel_file.html',
        file_info=file_info, variables=variables, data=data,
        subcontractor_data=subcontractor_data, state_data=state_data, district_data=district_data,
        block_data=block_data, errors=errors,
    )





# ----------------------------------  Save Excel file data in mysql tables --------------------------------------------------------------------------------------------------------------------------
# ✅ Route to Generate Excel
# Ensure the static/files directory exists
os.makedirs("static/files", exist_ok=True)


@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        headers = request.json.get("headers", [])
        data = request.json.get("data", [])

        if not headers or not data:
            return jsonify({"error": "Table data is empty!"}), 400

        df = pd.DataFrame(data, columns=headers)

        # ✅ Save the file on the server
        filename = "exported_data.xlsx"
        filepath = os.path.join("static/files", filename)
        df.to_excel(filepath, index=False)

        # ✅ Return JSON response with the file path
        return jsonify({"excel_path": f"/static/files/{filename}"})

    except Exception as e:
        print("🔴 ERROR:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/save_data', methods=['POST'])
def save_data():
    # Retrieve data from the request
    subcontractor_id = request.form.get("subcontractor_data")
    state_id = request.form.get("state_data")
    district_id = request.form.get("district_data")
    block_id = request.form.get("block_data")

    file_info = request.form.get('file_info')
    variables = request.form.getlist('variables')
    data = request.form.getlist('data')

    print('subcontractor_id :', subcontractor_id)
    print('state_id :', state_id)
    print('district_id :', district_id)
    print('block_id :', block_id)

    # Process the incoming data
    processed_data = [
        [0 if value in [None, 'None', ''] else value for value in row.split(',')]
        for row in data
    ]

    connection = mysql.connection.cursor()
    try:
        cursor = connection.cursor()

        for row in processed_data:
            invoice_id = None
            if row[2] in [0, '0', None, '']:
                row[2] = None

            words_in_row1 = row[1].lower().split()

            village_name = None
            work_type = None

            if 'village' in words_in_row1:
                village_position = words_in_row1.index('village')
                if village_position > 0:
                    village_name = " ".join(words_in_row1[:village_position])

            if 'work' in words_in_row1:
                work_position = words_in_row1.index('work')
                if work_position > 0:
                    work_type = " ".join(words_in_row1[village_position + 1: work_position + 1])

            if 'village' in words_in_row1 and 'work' in words_in_row1:
                village_id = None

                if block_id and village_name:
                    cursor.execute("""SELECT Village_Id FROM villages WHERE block_id = %s AND village_Name = %s
                        """, (block_id, village_name)
                                   )
                    return_village_id = cursor.fetchone()

                    if return_village_id:
                        village_id = return_village_id[0]
                    else:
                        cursor.execute(
                            """
                            INSERT INTO villages (Village_Name, Block_Id) VALUES (%s, %s)""",
                            (village_name, block_id)
                        )
                        cursor.execute("select village_id from villages order by village_id desc limit 1")
                        village_id = cursor.fetchone()[0]

                cursor.execute(
                    """
                    INSERT INTO Invoice (
                        PMC_No, Village_Id, Work_Type, Invoice_Details, Invoice_Date, Invoice_No,
                        Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount,
                        SD_Amount, On_Commission, Hydro_Testing, Hold_Amount, GST_SD_Amount, Final_Amount
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (row[0], int(village_id), work_type, row[1], row[2], row[3], row[4], row[5], row[6], row[8],
                          row[7], row[9], row[10], row[11], row[12], row[13], row[14], row[15]))
                # invoice_id = cursor.lastrowid

                cursor.execute("""
                    INSERT INTO payment (PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """, (row[0], row[3], row[16], row[17], row[18], row[19]))

                cursor.execute("INSERT INTO assing_subcontractors VALUES (%s, %s, %s) ",
                               (row[0], subcontractor_id, int(village_id)))

            elif 'gst' in words_in_row1 or 'release' in words_in_row1 or 'note' in words_in_row1:
                cursor.execute("""INSERT INTO gst_release (PMC_No, Invoice_No, Basic_Amount, Final_Amount)
                    VALUES (%s, %s, %s, %s)
                    """, (row[0], row[3], row[4], row[15]))

                cursor.execute("""INSERT INTO payment (PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """, (row[0], row[3], row[16], row[17], row[18], row[19]))

        connection.commit()
        return redirect(url_for('upload'))

    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        return jsonify({"error": "Database operation failed"}), 500
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "An error occurred"}), 500
    finally:
        cursor.close()
        connection.close()

def safe_decimal(value):
    """Convert None or empty values to 0.00 for decimal columns"""
    return float(value) if value not in [None, '', 'None'] else 0.00

# @app.route('/save_data', methods=['POST'])
# def save_data():
#     connection = None
#     cursor = None  # Ensure cursor is initialized before try block
#
#     try:
#         # Retrieve form data
#         subcontractor_id = request.form.get("subcontractor_data")
#         state_id = request.form.get("state_data")
#         district_id = request.form.get("district_data")
#         block_id = request.form.get("block_data")
#         table_data = request.form.get("table_data")
#
#         if not table_data:
#             return jsonify({"error": "No data received"}), 400
#
#         # Convert JSON string to list of lists
#         data = json.loads(table_data)
#
#         # Establish database connection
#         connection = mysql.connection.cursor()
#         cursor = connection.cursor(dictionary=True)
#
#         for row in data:
#             invoice_id = None
#             row[2] = None if row[2] in [0, '0', None, ''] else row[2]
#
#             words_in_row1 = row[1].lower().split()
#             village_name, work_type = None, None
#
#             if 'village' in words_in_row1:
#                 village_position = words_in_row1.index('village')
#                 if village_position > 0:
#                     village_name = " ".join(words_in_row1[:village_position])
#
#             if 'work' in words_in_row1:
#                 work_position = words_in_row1.index('work')
#                 if work_position > 0 and village_position is not None:
#                     work_type = " ".join(words_in_row1[village_position + 1: work_position + 1])
#
#             if village_name:
#                 cursor.execute("SELECT Village_Id FROM villages WHERE block_id = %s AND village_Name = %s",
#                                (block_id, village_name))
#                 village_result = cursor.fetchone()
#
#                 if village_result:
#                     village_id = village_result['Village_Id']
#                 else:
#                     cursor.execute("INSERT INTO villages (Village_Name, Block_Id) VALUES (%s, %s)",
#                                    (village_name, block_id))
#                     connection.commit()
#                     cursor.execute("SELECT Village_Id FROM villages ORDER BY Village_Id DESC LIMIT 1")
#                     village_id = cursor.fetchone()['Village_Id']
#
#                 # Update or Insert into Invoice table
#                 cursor.execute("""
#                     INSERT INTO Invoice (
#                         PMC_No, Village_Id, Work_Type, Invoice_Details, Invoice_Date, Invoice_No,
#                         Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount,
#                         SD_Amount, On_Commission, Hydro_Testing, Hold_Amount, GST_SD_Amount, Final_Amount
#                     ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
#                     ON DUPLICATE KEY UPDATE
#                         Work_Type=VALUES(Work_Type), Invoice_Details=VALUES(Invoice_Details),
#                         Invoice_Date=VALUES(Invoice_Date), Basic_Amount=VALUES(Basic_Amount),
#                         Debit_Amount=VALUES(Debit_Amount), After_Debit_Amount=VALUES(After_Debit_Amount),
#                         Amount=VALUES(Amount), GST_Amount=VALUES(GST_Amount), TDS_Amount=VALUES(TDS_Amount)
#                 """, (
#                     row[0], village_id, work_type, row[1], row[2], row[3],
#                     safe_decimal(row[4]), safe_decimal(row[5]), safe_decimal(row[6]), safe_decimal(row[8]),
#                     safe_decimal(row[7]), safe_decimal(row[9]), safe_decimal(row[10]), safe_decimal(row[11]),
#                     safe_decimal(row[12]), safe_decimal(row[13]), safe_decimal(row[14]), safe_decimal(row[15])
#                 ))
#
#                 # Update or Insert into Payment Table
#                 cursor.execute("""
#                     INSERT INTO payment (PMC_No, Invoice_No, Amount, TDS_Amount, Total_Amount, UTR)
#                     VALUES (%s, %s, %s, %s, %s, %s)
#                     ON DUPLICATE KEY UPDATE
#                         Amount=VALUES(Amount), TDS_Amount=VALUES(TDS_Amount), Total_Amount=VALUES(Total_Amount), UTR=VALUES(UTR)
#                 """, (
#                     row[0], row[3], safe_decimal(row[16]), safe_decimal(row[17]), safe_decimal(row[18]), row[19]
#                 ))
#                 # Updated code for gst_release
#                 cursor.execute("""
#                     INSERT INTO gst_release (PMC_No, Invoice_No, Basic_Amount, Final_Amount)
#                     VALUES (%s, %s, %s, %s)
#                     ON DUPLICATE KEY UPDATE
#                     Basic_Amount=VALUES(Basic_Amount),Final_Amount=VALUES(Final_Amount)
#                 """, (row[0], row[3],  safe_decimal(row[4]), safe_decimal(row[15])))
#
#         connection.commit()
#         return redirect(url_for('upload'))
#
#     except mysql.connector.Error as e:
#         print(f"Database error: {e}")
#         return jsonify({"error": f"Database operation failed: {e}"}), 500
#     except Exception as e:
#         print(f"Error: {e}")
#         return jsonify({"error": f"An error occurred: {e}"}), 500
#     finally:
#         # Safely close cursor and connection
#         if cursor:
#             cursor.close()
#         if connection:
#             connection.close()
#

# ------------------------------- Show Report Subcontractor  ---------------------
# @app.route('/report')
# def report_page():
#     return render_template('report.html')
#
#
# @app.route('/search_contractor', methods=['POST'])
# def search_contractor():
#     connection = get_db_connection()
#     cursor = connection.cursor(dictionary=True)
#     subcontractor_name = request.form.get('subcontractor_name')
#     query = """
#         SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
#                         i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
#                         i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
#                         i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
#         FROM subcontractors s
#         LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#         LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#         LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
#         WHERE s.Contractor_Name = %s order by i.PMC_No;
#     """
#     cursor.execute(query, (subcontractor_name,))
#     data = cursor.fetchall()
#     return jsonify(data)
#
#
# @app.route('/download')
# def download_file():
#     connection = get_db_connection()
#     output_folder = "static/download"
#     output_file = os.path.join(output_folder, "Subcontractor_Report.xlsx")
#
#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)
#
#     if connection:
#         cursor = connection.cursor(dictionary=True)
#         query = """SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
#                             i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
#                             i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
#                             i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
#                    FROM subcontractors s
#                    LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#                    LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#                    LEFT JOIN invoice i ON v.Village_Id = i.Village_Id  order by i.PMC_No;"""
#         cursor.execute(query)
#         rows = cursor.fetchall()
#         columns = [desc[0] for desc in cursor.description]
#
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.append(columns)
#         for row in rows:
#             sheet.append([row[column] for column in columns])
#         workbook.save(output_file)
#
#     return send_from_directory(output_folder, "Subcontractor_Report.xlsx", as_attachment=True)

#  end searching controller....


@app.route('/report')
def report_page():
    return render_template('report.html')


# Search list
@app.route('/search_contractor', methods=['POST'])
def search_contractor():
    connection = mysql.connection.cursor()
    cursor = connection.cursor(dictionary=True)

    subcontractor_name = request.form.get('subcontractor_name')
    pmc_no = request.form.get('pmc_no')
    state = request.form.get('state')
    district = request.form.get('district')
    block = request.form.get('block')
    village = request.form.get('village')
    year_from = request.form.get('year_from')
    year_to = request.form.get('year_to')

    conditions = []
    params = []

    if subcontractor_name:
        conditions.append("LOWER(s.Contractor_Name) LIKE LOWER(%s)")
        params.append(f"%{subcontractor_name}%")
    if pmc_no:
        conditions.append("i.PMC_No = %s")
        params.append(pmc_no)
    if state:
        conditions.append("LOWER(st.State_Name) LIKE LOWER(%s)")
        params.append(f"%{state}%")
    if district:
        conditions.append("LOWER(d.District_Name) LIKE LOWER(%s)")
        params.append(f"%{district}%")
    if block:
        conditions.append("LOWER(b.Block_Name) LIKE LOWER(%s)")
        params.append(f"%{block}%")
    if village:
        conditions.append("LOWER(v.Village_Name) LIKE LOWER(%s)")
        params.append(f"%{village}%")
    if year_from and year_to:
        conditions.append("i.Invoice_Date BETWEEN %s AND %s")
        params.append(year_from)
        params.append(year_to)

    if not conditions:
        return jsonify({"error": "At least one field is required for search."}), 400

    query = f""" SELECT DISTINCT s.Contractor_Id, s.Contractor_Name, i.PMC_No, st.State_Name, 
        d.District_Name, b.Block_Name, v.Village_Name 
        FROM subcontractors s
        LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
        LEFT JOIN districts d ON b.District_id = d.District_id
        LEFT JOIN states st ON d.State_Id = st.State_Id
        LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
        WHERE {' AND '.join(conditions)}
    """

    cursor.execute(query, tuple(params))
    data = cursor.fetchall()
    return jsonify(data)


#  genete report by contractor Id --------------------------
@app.route('/contractor_report/<int:contractor_id>')
def contractor_report(contractor_id):
    connection = mysql.connection.cursor()
    cursor = connection.cursor(dictionary=True)
    contractor_id = contractor_id

    cursor.execute(""" SELECT DISTINCT s.Contractor_Name,st.State_Name,d.District_Name, b.Block_Name,
    s.Mobile_No,s.GST_Registration_Type,s.GST_No
    FROM subcontractors s
    LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
    LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
    LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
    LEFT JOIN districts d ON b.District_id = d.District_id
    LEFT JOIN states st ON d.State_Id = st.State_Id
    WHERE s.Contractor_Id = %s""", (contractor_id,))
    contInfo = cursor.fetchone()

    query = """
        SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
                        i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
                        i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
                        i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
        FROM subcontractors s
        LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
        WHERE s.Contractor_Id = %s
    """

    cursor.execute(query, (contractor_id,))
    invoices = cursor.fetchall()

    return render_template('subcontractor_report.html', invoices=invoices, contInfo=contInfo,
                           contractor_id=contractor_id)


#  Download report by contractor Id -----------------------
# @app.route('/download_report/<int:contractor_id>')
# def download_report(contractor_id):
#     connection = get_db_connection()
#     output_folder = "static/download"
#     output_file = os.path.join(output_folder, f"Contractor_Report_{contractor_id}.xlsx")
#
#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)
#
#     cursor = connection.cursor(dictionary=True)
#
#     cursor.execute(""" SELECT DISTINCT s.Contractor_Name,st.State_Name,d.District_Name, b.Block_Name
#         FROM subcontractors s
#         LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#         LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#         LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
#         LEFT JOIN districts d ON b.District_id = d.District_id
#         LEFT JOIN states st ON d.State_Id = st.State_Id
#         WHERE s.Contractor_Id = %s""", (contractor_id,))
#     contInfo = cursor.fetchone()
#
#     query = """
#         SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
#                         i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
#                         i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
#                         i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
#         FROM subcontractors s
#         LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#         LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#         LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
#         WHERE s.Contractor_Id = %s
#     """
#
#     cursor.execute(query, (contractor_id,))
#     rows = cursor.fetchall()
#     columns = [desc[0] for desc in cursor.description]
#
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.append(columns)
#
#     for row in rows:
#         sheet.append([row[column] for column in columns])
#
#     workbook.save(output_file)
#
#     return send_from_directory(output_folder, f"Contractor_Report_{contractor_id}.xlsx", as_attachment=True)

@app.route('/download_report/<int:contractor_id>')
def download_report(contractor_id):
    connection = mysql.connection.cursor()
    cursor = connection.cursor(dictionary=True)

    output_folder = "static/download"
    output_file = os.path.join(output_folder, f"Contractor_Report_{contractor_id}.xlsx")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    contractor_query = """
        SELECT DISTINCT s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
                        s.Mobile_No, s.GST_Registration_Type, s.GST_No
        FROM subcontractors s
        LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
        LEFT JOIN districts d ON b.District_id = d.District_id
        LEFT JOIN states st ON d.State_Id = st.State_Id
        WHERE s.Contractor_Id = %s
    """
    cursor.execute(contractor_query, (contractor_id,))
    contInfo = cursor.fetchone()

    if not contInfo:
        return "No contractor found", 404

    # Fetch invoice details
    invoice_query = """
        SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
                        i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
                        i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
                        i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
        FROM subcontractors s
        LEFT JOIN assing_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
        WHERE s.Contractor_Id = %s
    """
    cursor.execute(invoice_query, (contractor_id,))
    invoices = cursor.fetchall()

    # Create Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Contractor Report"

    # Add contractor details
    sheet.append(["Contractor Name", contInfo["Contractor_Name"]])
    sheet.append(["State", contInfo["State_Name"]])
    sheet.append(["District", contInfo["District_Name"]])
    sheet.append(["Block", contInfo["Block_Name"]])

    # Add invoice details if available
    if invoices:
        sheet.append(["PMC No", "Village Name", "Work Type", "Invoice Details", "Invoice Date",
                      "Invoice No", "Basic Amount", "Debit Amount", "After Debit Amount",
                      "Amount", "GST Amount", "TDS Amount", "SD Amount", "On Commission",
                      "Hydro Testing", "Hold Amount", "GST SD Amount", "Final Amount"])

        for row in invoices:
            sheet.append(
                [row["PMC_No"], row["Village_Name"], row["Work_Type"], row["Invoice_Details"], row["Invoice_Date"],
                 row["Invoice_No"], row["Basic_Amount"], row["Debit_Amount"], row["After_Debit_Amount"],
                 row["Amount"], row["GST_Amount"], row["TDS_Amount"], row["SD_Amount"], row["On_Commission"],
                 row["Hydro_Testing"], row["Hold_Amount"], row["GST_SD_Amount"], row["Final_Amount"]])

    workbook.save(output_file)

    cursor.close()
    connection.close()

    return send_from_directory(output_folder, f"Contractor_Report_{contractor_id}.xlsx", as_attachment=True)


# Generate report by PMC No -----------------------
# @app.route('/pmc_report/<pmc_no>')
# def pmc_report(pmc_no):
#     connection = mysql.connection.cursor()
#     cursor = connection.cursor(dictionary=True)
#
#     query = """
#         SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
#                         i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
#                         i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
#                         i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
#         FROM invoice i
#         LEFT JOIN villages v ON i.Village_Id = v.Village_Id
#         WHERE i.PMC_No = %s
#     """
#
#     cursor.execute(query, (pmc_no,))
#     invoices = cursor.fetchall()
#
#     return render_template('pmc_report.html', invoices=invoices, pmc_no=pmc_no)
#
#
# # Download report by PMC No
# @app.route('/download_pmc_report/<pmc_no>')
# def download_pmc_report(pmc_no):
#     connection = mysql.connection.cursor()
#     output_folder = "static/download"
#     output_file = os.path.join(output_folder, f"PMC_Report_{pmc_no}.xlsx")
#
#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)
#
#     cursor = connection.cursor(dictionary=True)
#     query = """
#         SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details, i.Invoice_Date,
#                         i.Invoice_No, i.Basic_Amount, i.Debit_Amount, i.After_Debit_Amount,
#                         i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount, i.On_Commission,
#                         i.Hydro_Testing, i.Hold_Amount, i.GST_SD_Amount, i.Final_Amount
#         FROM invoice i
#         LEFT JOIN villages v ON i.Village_Id = v.Village_Id
#         WHERE i.PMC_No = %s
#     """
#
#     cursor.execute(query, (pmc_no,))
#     rows = cursor.fetchall()
#     columns = [desc[0] for desc in cursor.description]
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.append(columns)
#
#     for row in rows:
#         sheet.append([row[column] for column in columns])
#
#     workbook.save(output_file)
#
#     return send_from_directory(output_folder, f"PMC_Report_{pmc_no}.xlsx", as_attachment=True)
@app.route('/pmc_report/<pmc_no>')
def pmc_report(pmc_no):
    connection = mysql.connection.cursor()
    cursor = connection.cursor(dictionary=True)

    query = """
        SELECT DISTINCT 
    i.PMC_No, 
    v.Village_Name, 
    i.Work_Type, 
    i.Invoice_Details, 
    i.Invoice_Date,
    i.Invoice_No, 
    i.Basic_Amount, 
    i.Debit_Amount, 
    i.After_Debit_Amount,
    i.Amount, 
    i.GST_Amount, 
    i.TDS_Amount, 
    i.SD_Amount, 
    i.On_Commission,
    i.Hydro_Testing, 
    i.Hold_Amount, 
    i.GST_SD_Amount, 
    i.Final_Amount,
    gr.Basic_Amount AS GST_Release_Amount, 
    gr.Final_Amount AS GST_Release_Final_Amount,
    p.Amount AS Payment_Amount,
    p.TDS_Amount AS Payment_TDS_Amount,
    p.Total_Amount AS Payment_Total_Amount,
    p.UTR AS Payment_UTR
FROM 
    invoice i
LEFT JOIN 
    villages v ON i.Village_Id = v.Village_Id
LEFT JOIN 
    gst_release gr ON i.PMC_No = gr.PMC_No AND i.Invoice_No = gr.Invoice_No
LEFT JOIN 
    payment p ON i.PMC_No = p.PMC_No AND i.Invoice_No = p.Invoice_No
WHERE 
    i.PMC_No = %s

    """

    cursor.execute(query, (pmc_no,))
    invoices = cursor.fetchall()

    return render_template('pmc_report.html', invoices=invoices, pmc_no=pmc_no)


# Download report by PMC No
@app.route('/download_pmc_report/<pmc_no>')
def download_pmc_report(pmc_no):
    connection = mysql.connection.cursor()
    output_folder = "static/download"
    output_file = os.path.join(output_folder, f"PMC_Report_{pmc_no}.xlsx")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    cursor = connection.cursor(dictionary=True)
    query = """
          SELECT DISTINCT 
    i.PMC_No, 
    v.Village_Name, 
    i.Work_Type, 
    i.Invoice_Details, 
    i.Invoice_Date,
    i.Invoice_No, 
    i.Basic_Amount, 
    i.Debit_Amount, 
    i.After_Debit_Amount,
    i.Amount, 
    i.GST_Amount, 
    i.TDS_Amount, 
    i.SD_Amount, 
    i.On_Commission,
    i.Hydro_Testing, 
    i.Hold_Amount, 
    i.GST_SD_Amount, 
    i.Final_Amount,
    gr.Basic_Amount AS GST_Release_Amount, 
    gr.Final_Amount AS GST_Release_Final_Amount,
    p.Amount AS Payment_Amount,
    p.TDS_Amount AS Payment_TDS_Amount,
    p.Total_Amount AS Payment_Total_Amount,
    p.UTR AS Payment_UTR
FROM 
    invoice i
LEFT JOIN 
    villages v ON i.Village_Id = v.Village_Id
LEFT JOIN 
    gst_release gr ON i.PMC_No = gr.PMC_No AND i.Invoice_No = gr.Invoice_No
LEFT JOIN 
    payment p ON i.PMC_No = p.PMC_No AND i.Invoice_No = p.Invoice_No
WHERE 
    i.PMC_No = %s
    """

    cursor.execute(query, (pmc_no,))
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(columns)

    for row in rows:
        sheet.append([row[column] for column in columns])

    workbook.save(output_file)

    return send_from_directory(output_folder, f"PMC_Report_{pmc_no}.xlsx", as_attachment=True)



if __name__ == '__main__':
    #config_db()
    app.run(host='0.0.0.0', port=5000, debug=True)

